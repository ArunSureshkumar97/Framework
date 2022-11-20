import openpyxl
from openpyxl.styles import PatternFill

def getRow(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)

def getColumn(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_column)

def Readdata(file,sheetName,rownum,columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum,columno).value

def writedata(file,sheetName,rownum,columno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum,columno).value = data
    workbook.save(file)

# def fillGreencolor(file,sheetName,rownum,columno):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook[sheetName]
#     greenFill = PatternFill(start_color='60b212',end_color='60b212,fill_type='solid')
#     sheet.cell(rownum, columno).fill = greenFill
#     workbook.save(file)
#
# def fillRedcolor(file,sheetName,rownum,columno):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook[sheetName]
#     redFill = PatternFill(start_color='ff0000',end_color='ff0000,fill_type='solid')
#     sheet.cell(rownum, columno).fill = redFill
#     workbook.save(file)



